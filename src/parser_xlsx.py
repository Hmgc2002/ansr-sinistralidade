"""Normalizes the ANSR 'Anexo' xlsx workbooks into structured CSV output.

Two outputs:
1. Raw dump: every sheet of every annual workbook, exported verbatim as CSV
   (data/processed/xlsx_raw/<year>/<sheet_id>.csv), plus a master index
   (data/processed/xlsx_tables_index.csv) mapping year/sheet_id -> table
   title, taken from each workbook's own 'Índice' sheet.
2. Tidy time series, in two separate files because they are not the same
   measurement:
   - data/processed/sinistralidade_mensal.csv: the "Sinistralidade em
     Portugal por mês" table (whole country, "30 dias" fatality count)
     from every annual workbook (2020-2024) and the 2025 monthly report.
   - data/processed/sinistralidade_mensal_continente_24h.csv: the
     "Sinistralidade no Continente por mês" table found in the "24h"
     annual workbooks (2023-2024) — mainland only (excludes Açores/
     Madeira) AND a different fatality-count window (24h vs 30 dias).
   Both have columns: report_year, scope, month, month_num,
   acidentes_com_vitimas, vitimas_mortais, feridos_graves, feridos_leves.

Note: some workbooks (the 2025 monthly report, and the "24h" annual
reports from 2023+) use a simpler sheet layout that combines what used
to be separate Quadros into one sheet (e.g. '4 e 5' instead of separate
'4' and '5' tabs) — `_find_monthly_sheet` handles this, but it's only
been validated against the handful of files seen so far.

Usage:
    python src/parser_xlsx.py
"""
from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "data" / "raw"
PROCESSED_DIR = ROOT / "data" / "processed"
XLSX_RAW_OUT = PROCESSED_DIR / "xlsx_raw"
INDEX_OUT = PROCESSED_DIR / "xlsx_tables_index.csv"
MONTHLY_OUT = PROCESSED_DIR / "sinistralidade_mensal.csv"
CONTINENTE_OUT = PROCESSED_DIR / "sinistralidade_mensal_continente_24h.csv"

MONTHS_PT = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}
GROUPS = ["AcV", "VM", "FG", "FL"]


def _safe(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', "_", name).strip()


def find_index_sheet(wb: openpyxl.Workbook) -> str | None:
    for name in wb.sheetnames:
        if name.strip().lower() in ("índice", "indice"):
            return name
    return None


def read_index(wb: openpyxl.Workbook) -> list[tuple[str, str]]:
    """Returns [(quadro_number, title), ...] parsed from the Índice sheet."""
    idx_name = find_index_sheet(wb)
    if not idx_name:
        return []
    ws = wb[idx_name]
    out = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str) and v.strip().lower().startswith("quadro"):
                m = re.match(r"quadro\s+([\d.]+)\.?\s*(.*)", v.strip(), re.IGNORECASE)
                if m:
                    out.append((m.group(1).rstrip("."), m.group(2).strip()))
    return out


def dump_raw_sheets(xlsx_path: Path, year: str, index_rows: list[dict]) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    titles = dict(read_index(wb))
    out_dir = XLSX_RAW_OUT / year
    out_dir.mkdir(parents=True, exist_ok=True)

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in ("índice", "indice", "siglas"):
            continue
        ws = wb[sheet_name]
        title = titles.get(sheet_name.strip(), "")
        safe_sheet = _safe(sheet_name)
        out_path = out_dir / f"{safe_sheet}.csv"
        with out_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
        index_rows.append(
            {
                "year": year,
                "source_file": xlsx_path.name,
                "sheet_id": sheet_name.strip(),
                "title": title,
                "csv_path": str(out_path.relative_to(ROOT)),
                "rows": ws.max_row,
                "cols": ws.max_column,
            }
        )
    wb.close()


def _find_monthly_sheet(wb: openpyxl.Workbook, scope_phrase: str) -> str | None:
    for quadro_num, title in read_index(wb):
        t = title.lower()
        if "por mês" in t and scope_phrase in t and "variação" not in t:
            if quadro_num in wb.sheetnames:
                return quadro_num
            # some workbooks (e.g. the 2025 monthly report, and the "24h"
            # annual reports) combine several "Quadro N" entries into one
            # sheet named like "4 e 5" — match quadro_num against the sheet
            # name's tokens instead of exact
            for sheet_name in wb.sheetnames:
                if quadro_num in re.split(r"\W+", sheet_name):
                    return sheet_name
    return None


def extract_monthly_series(xlsx_path: Path, report_year: int, scope_phrase: str, scope_label: str) -> list[dict]:
    """scope_phrase selects the sheet ("sinistralidade em portugal" for the
    "30 dias" reports, "sinistralidade no continente" for the "24h" ones —
    these are two different geographic scopes, not just two fatality-count
    methodologies: Continente excludes the Azores/Madeira). scope_label is
    recorded on each output row so the two series are never conflated.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet_id = _find_monthly_sheet(wb, scope_phrase)
    if not sheet_id:
        wb.close()
        return []

    ws = wb[sheet_id]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_row_idx = next(
        (i for i, r in enumerate(rows) if any(isinstance(v, str) and v.strip() == "AcV" for v in r)),
        None,
    )
    if header_row_idx is None:
        return []
    group_row = rows[header_row_idx]
    year_row = rows[header_row_idx + 1]

    group_starts = [(i, v.strip()) for i, v in enumerate(group_row) if isinstance(v, str) and v.strip() in GROUPS]
    col_for_group: dict[str, int] = {}
    for gi, (start_col, label) in enumerate(group_starts):
        end_col = group_starts[gi + 1][0] if gi + 1 < len(group_starts) else len(group_row)
        target_col = None
        for c in range(start_col, end_col):
            if year_row[c] == report_year:
                target_col = c
                break
        if target_col is not None:
            col_for_group[label] = target_col

    if len(col_for_group) != len(GROUPS):
        return []

    out = []
    for row in rows[header_row_idx + 2:]:
        month_label = row[0]
        if not isinstance(month_label, str):
            continue
        key = month_label.strip().lower()[:3]
        if key == "tot":
            break
        month_num = MONTHS_PT.get(key)
        if month_num is None:
            continue
        out.append(
            {
                "report_year": report_year,
                "scope": scope_label,
                "month": month_label.strip(),
                "month_num": month_num,
                "acidentes_com_vitimas": row[col_for_group["AcV"]],
                "vitimas_mortais": row[col_for_group["VM"]],
                "feridos_graves": row[col_for_group["FG"]],
                "feridos_leves": row[col_for_group["FL"]],
            }
        )
    return out


def main() -> None:
    xlsx_files = sorted(RAW_DIR.glob("*/*.xlsx"))
    if not xlsx_files:
        print("Nenhum .xlsx encontrado em data/raw — corre primeiro downloader.py", file=sys.stderr)
        sys.exit(1)

    index_rows: list[dict] = []
    monthly_rows: list[dict] = []
    continente_rows: list[dict] = []

    for path in xlsx_files:
        year = path.parent.name
        print(f"A processar {year}: {path.name}")
        dump_raw_sheets(path, year, index_rows)

        if year.isdigit() and 2020 <= int(year) <= 2025:
            extracted = extract_monthly_series(path, int(year), "sinistralidade em portugal", "Portugal")
            if extracted:
                monthly_rows.extend(extracted)
                print(f"  -> {len(extracted)} meses (Portugal) extraídos para a série nacional")

            extracted_cont = extract_monthly_series(path, int(year), "sinistralidade no continente", "Continente")
            if extracted_cont:
                continente_rows.extend(extracted_cont)
                print(f"  -> {len(extracted_cont)} meses (Continente, 24h) extraídos")

            if not extracted and not extracted_cont:
                print(f"  -> aviso: não encontrei a tabela mensal neste ficheiro")

    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    with INDEX_OUT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(index_rows[0].keys()))
        writer.writeheader()
        writer.writerows(index_rows)
    print(f"\n{len(index_rows)} tabelas catalogadas -> {INDEX_OUT}")

    if monthly_rows:
        monthly_rows.sort(key=lambda r: (r["report_year"], r["month_num"]))
        with MONTHLY_OUT.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(monthly_rows[0].keys()))
            writer.writeheader()
            writer.writerows(monthly_rows)
        print(f"{len(monthly_rows)} linhas na série mensal nacional (Portugal, 30 dias) -> {MONTHLY_OUT}")
    else:
        print("Aviso: nenhuma linha extraída para a série mensal nacional.", file=sys.stderr)

    if continente_rows:
        continente_rows.sort(key=lambda r: (r["report_year"], r["month_num"]))
        with CONTINENTE_OUT.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(continente_rows[0].keys()))
            writer.writeheader()
            writer.writerows(continente_rows)
        print(f"{len(continente_rows)} linhas na série mensal Continente (24h) -> {CONTINENTE_OUT}")


if __name__ == "__main__":
    main()
