"""Extracts the annual summary table from the ANSR "PN-ISSR Recomendações"
Excel workbook — a small, different-granularity companion to
pontos_negros.csv. That CSV lists individual black-spot road segments
(2019-2022, one row per stretch of road); this one is national yearly
*counts* (how many black spots identified, how many inspections, how
many recommendations issued/implemented) — already includes 2023, which
the per-segment PDF extraction doesn't (no "PN 2023" detailed PDF exists
at this source, confirmed by a live re-check — see README).

The workbook is republished monthly with the current month baked into
both the filename ("PN <MÊS>. <ANO>.xlsx") and the sheet's title cell
("Situação em <mês> de <ano>") — so the exact filename isn't stable
across runs, and is discovered here from the source page's link list by
extension instead of a hardcoded name.

Usage:
    python src/parser_pontos_negros_resumo.py
Writes:
    data/processed/pontos_negros_resumo_anual.csv
"""
from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path

import csv
import openpyxl
import requests
from bs4 import BeautifulSoup

ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = ROOT / "data" / "processed" / "pontos_negros_resumo_anual.csv"

SOURCE_PAGE = "http://www.ansr.pt/SegurancaRodoviaria/PontosNegrosRecomendacoes/Pages/default.aspx"
TIMEOUT = 60
USER_AGENT = (
    "ansr-sinistralidade-scraper/0.1 "
    "(uso pessoal/pesquisa; contacto: hmgc2016@proton.me)"
)

COLUMNS = [
    "n_pn", "pn_recorrentes", "n_issr_pn", "n_issr_outros", "n_issr_total",
    "recomendacoes_emitidas", "recomendacoes_implementadas", "taxa_execucao_pct",
]


def find_workbook_url() -> str | None:
    resp = requests.get(SOURCE_PAGE, timeout=TIMEOUT, headers={"User-Agent": USER_AGENT})
    resp.raise_for_status()
    resp.encoding = resp.apparent_encoding or "utf-8"
    soup = BeautifulSoup(resp.text, "html.parser")
    candidates = [
        a["href"] for a in soup.select("a[href]")
        if a["href"].lower().endswith(".xlsx") and re.search(r"^PN[\s._-]", a.get_text(strip=True), re.IGNORECASE)
    ]
    return candidates[0] if candidates else None


def parse_summary_sheet(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = next(i for i, r in enumerate(rows) if r and r[2] == "Anos")
    data_start = header_idx + 1

    out = []
    total_row = None
    for row in rows[data_start:]:
        label = row[2]
        if label is None:
            continue
        if isinstance(label, str) and label.strip().lower().startswith("total"):
            total_row = row
            break
        ano = str(label).rstrip("*").strip()
        values = row[3:3 + len(COLUMNS)]
        record = {"ano": ano}
        for col, v in zip(COLUMNS, values):
            record[col] = None if v in (None, "-") else v
        if record["taxa_execucao_pct"] is not None:
            record["taxa_execucao_pct"] = round(record["taxa_execucao_pct"] * 100, 1)
        out.append(record)

    if total_row is not None:
        # sanity check: the source's own "Total" row should equal the sum
        # of the per-year rows for every column that has no gaps (2023's
        # "-" columns make a full-column sum meaningless, so those are
        # skipped rather than compared)
        for i, col in enumerate(COLUMNS[:-1]):  # all but the %, which isn't additive
            values = [r[col] for r in out if r[col] is not None]
            if len(values) == len(out):  # only compare if no year had a gap in this column
                computed = sum(values)
                sourced = total_row[3 + i]
                if computed != sourced:
                    print(f"aviso: total de '{col}' não bate — calculado {computed}, fonte diz {sourced}")

    return out


def main() -> None:
    url = find_workbook_url()
    if not url:
        print("Aviso: não encontrei nenhum .xlsx 'PN ...' na página fonte — a estrutura pode ter mudado.")
        return
    print(f"A descarregar {url} ...")
    resp = requests.get(url, timeout=TIMEOUT, headers={"User-Agent": USER_AGENT})
    resp.raise_for_status()
    wb = openpyxl.load_workbook(BytesIO(resp.content), data_only=True)

    rows = parse_summary_sheet(wb)
    if not rows:
        print("Nenhuma linha extraída.")
        return

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUT_PATH.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["ano"] + COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"{len(rows)} anos ({rows[0]['ano']}-{rows[-1]['ano']}) -> {OUT_PATH}")


if __name__ == "__main__":
    main()
